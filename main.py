from PyQt6.QtWidgets import QApplication, QMainWindow, QGraphicsDropShadowEffect, QLabel, QHBoxLayout, QVBoxLayout, QPushButton, QFrame, QLineEdit, QSizePolicy
from PyQt6.QtGui import QPixmap, QImage, QFont
from PyQt6.QtCore import pyqtSignal, pyqtSlot, Qt, QThread
from PyQt6 import uic
from datetime import datetime
from ultralytics import YOLO
from ultralytics.utils.plotting import Annotator
from openpyxl import Workbook, load_workbook
import cv2, numpy, pandas


datapath = r"file_absensi/data_absensi.xlsx"
days = ["Senin", "Selasa", "Rabu", "Kamis", "Jumat", "Sabtu", "Minggu"]
months = ["Januari", "Februari", "Maret", "April", "Mei", "Juni", "Juli", "Agustus", "September", "Oktober", "November", "Desember"]
names = {"Foo":1,"Rizal Fuadi":2,"Komeri":3,"Hari Solikah":4,"Waryanto":5,"Choiriyaningrum":6,"Agus Subandono":7,"Fathan Sholeh":8,"Danang Windriyo":9,"Marji":10,"Sapto":11,"Sarah Octavia":12,"Mahendro":13,"Kukuh Ega Prastya":14,"Nana Romatul":15}


class videoThread(QThread):
    model = YOLO("resource/best-new-10.pt")
    pixmapSignal = pyqtSignal(numpy.ndarray)
    detectSignal = pyqtSignal(str)

    def __init__(self):
        super().__init__()
        self._run_flag = True

        # create webcam capturer
        self.captr = cv2.VideoCapture(0)
        self.captr.set(3, 640)
        self.captr.set(4, 360)

    def run(self):
        while self._run_flag:
            success, capimg = self.captr.read()
            
            # process image to model
            res = self.model.predict(capimg, imgsz=[288, 512], verbose=False, vid_stride=True, max_det=6, conf=0.5)

            # draw bounding box
            anotator = Annotator(capimg)
            for r in res:
                boxes = r.boxes
                for box in boxes:
                    xy = box.xyxy[0]
                    cf = float(box.conf)
                    dn = self.model.names[int(box.cls)]

                    anotator.box_label(xy, dn+"-"+str(round(cf, 2)))
                    # submit attendance if confidence > 0.85
                    if(cf >= 0.85): 
                        self.detectSignal.emit(dn)
                    print(dn, " ", cf)

            resimg = anotator.result()

            # passing out img
            if success:
                self.pixmapSignal.emit(resimg)

    def stop(self):
        self._run_flag = False
        self.wait()


class manualForm(QFrame):
    def __init__(self):
        super().__init__()

        manualContainer = QVBoxLayout()
        inputName = QLineEdit("Nama")
        manualContainer.addWidget(inputName)

        self.setLayout(manualContainer)


class widgetList(QVBoxLayout):
    def deleter(self):
        self.hide()

    def __init__(self, name, keterangan, date):
        super().__init__()

        lfont = QFont("Times", 12)
        label_name = QLabel(name)
        label_name.setFont(lfont)
        self.addWidget(label_name)

        label_ket = QLabel(keterangan)
        label_date = QLabel(date, alignment=Qt.AlignmentFlag.AlignRight)

        botContainer = QHBoxLayout()
        botContainer.addWidget(label_ket)
        botContainer.addWidget(label_date)

        # btn_delete = QPushButton("Hapus")
        # btn_delete.clicked.connect(self.deleter)
        # botContainer.addWidget(btn_delete)

        self.addLayout(botContainer)


class UI(QMainWindow):
    def __init__(self):
        super().__init__()

        # loading the ui file with uic module
        uic.loadUi("resource/absen_main.ui", self)
        
        self.updateDate()
        self.loadAbsenData()
            
        # list to check if names ever submitted
        self.listKehadiran = []

        # save starting minute used as comparison to update date label
        self.saved_minute = self.minute
        self.label_clock.setText(self.time)
        self.label_date.setText(self.day+", "+self.date)

        self.frameList.setLayout(self.vlListBigContainer)
        self.toggleManualIinput = False
        self.frameInput.setLayout(self.vlInputLeft)
        self.frameInput.hide()
        self.framePetunjuk.setLayout(self.vlPetunjuk)
        self.framePetunjuk.hide()
        
        # thread
        self.thread = videoThread()
        self.thread.pixmapSignal.connect(self.updateVideo)
        self.thread.detectSignal.connect(self.submitDetected)
        self.thread.start()

        # connect button to function
        self.btn_petunjuk.clicked.connect(self.petunjukFunc)
        self.btn_daftarMasuk.clicked.connect(self.daftarMasukFunc)
        self.btn_inputKedatangan.clicked.connect(self.inputFunc)
        self.btn_submit_manual.clicked.connect(self.btnManualInput)

        # create shadow obj
        shadowLogo  = QGraphicsDropShadowEffect() 
        shadowTittle  = QGraphicsDropShadowEffect() 
        shadowClock  = QGraphicsDropShadowEffect() 
        shadowDate = QGraphicsDropShadowEffect() 
  
        # setting blur radius
        shadowLogo.setBlurRadius(10)
        shadowTittle.setBlurRadius(20)
        shadowClock.setBlurRadius(30)
        shadowDate.setBlurRadius(30)
  
        # adding shadow to the label
        self.mainLogo.setGraphicsEffect(shadowLogo)
        self.label_tittle.setGraphicsEffect(shadowTittle)
        self.label_clock.setGraphicsEffect(shadowClock)
        self.label_date.setGraphicsEffect(shadowDate)
        

    def inputFunc(self):

        if self.toggleManualIinput:
            self.frameInput.hide()
            self.cbName.setCurrentIndex(0)
            self.cbPresence.setCurrentIndex(0)
            self.btn_inputKedatangan.setText("  Input Kedatangan  ")
            self.toggleManualIinput = False
        else:
            self.frameInput.show()
            self.btn_inputKedatangan.setText("  Tutup Input Kedatangan  ")
            self.toggleManualIinput = True

    def btnManualInput(self):
        self.submitKedatangan(self.cbName.currentText(), self.cbPresence.currentText())

    def submitDetected(self, name):
        self.submitKedatangan(name, "Hadir")

    def submitKedatangan(self, sname, spresence):
        if not sname in self.listKehadiran:
            # add name to list 
            nameItem = widgetList(sname, spresence, self.fulldate)
            self.vlListMain.insertLayout(0, nameItem)
            self.listKehadiran.append(sname)

            # change presence in data table
            writer = pandas.ExcelWriter(datapath, engine="openpyxl", mode='a', if_sheet_exists="overlay")
            self.dataKehadiran.loc[int(names[sname]-1), int(self.sdate)] = spresence
            self.dataKehadiran.to_excel(writer, index=False, sheet_name=self.sheetname)
            writer.close()

            #print(self.dataKehadiran)

    def updateDate(self):
        now = datetime.now()
        today = datetime.today()
        
        self.minute = now.minute
        self.sdate = now.day
        self.month = now.month
        self.year = now.year
        self.time = now.strftime("%H:%M")
        self.day = days[today.weekday()]
        self.date = today.strftime("%d/%m/%Y")
        self.fulldate = self.time+"/"+self.date
        self.sheetname = months[self.month-1]+str(self.year)

    def loadAbsenData(self):
        try:
            self.dataKehadiran = pandas.read_excel(datapath, sheet_name=self.sheetname)
        except FileNotFoundError:
            '''create new data and write name/date'''
            self.wb = Workbook()
            self.absenDataCreate()
        except ValueError:
            '''create new sheet and write name/date'''
            self.wb = load_workbook(datapath)
            self.absenDataCreate()

        self.dataKehadiran = pandas.read_excel(datapath, sheet_name=self.sheetname)

    def absenDataCreate(self):
        ws = self.wb.create_sheet(self.sheetname)
        ws.cell(row=1, column=1).value = "nama/tanggal"

        '''
        Sheet Format

        nama/tanggal | 1 | 2 | 3 | ... | 31
        member-1     | H | T | I |     |  
        ....         |   |   |   |     |
        member-n     | H | T | I |     |   
        
        H Hadir, T Tidak Hadir, I Izin
        '''

        # create columns for day 1 to 31
        for i in range(1,32):
            ws.cell(row=1, column=i+1).value = i
        # write all available name from data to rows
        for n in list(names.items()):
            ws.cell(row=n[1]+1, column=1).value = n[0]

        self.wb.save(datapath)

    def petunjukFunc(self):
        self.framePetunjuk.show()
        self.frameList.hide()

    def daftarMasukFunc(self):
        self.frameList.show()
        self.framePetunjuk.hide()

    @pyqtSlot(numpy.ndarray)
    def updateVideo(self, img):

        img = self.cvImgConvert(img)
        self.labelCamera.setPixmap(img)

        # update time each minute
        self.updateDate()
        if self.minute != self.saved_minute:
            self.saved_minute = self.minute
            self.label_clock.setText(self.time)

    def cvImgConvert(self, img_to_process):
        """Convert from an opencv image to QPixmap"""
        rgb_image = cv2.cvtColor(img_to_process, cv2.COLOR_BGR2RGB)
        h, w, ch = rgb_image.shape
        bytes_per_line = ch * w
        convert_to_Qt_format = QImage(rgb_image.data, w, h, bytes_per_line, QImage.Format.Format_RGB888)
        res_img = convert_to_Qt_format.scaled(1280, 720, Qt.AspectRatioMode.IgnoreAspectRatio)
        return QPixmap.fromImage(res_img)

    # close thread when app closed
    def closeEvent(self, event):
        self.thread.stop()
        event.accept()


app = QApplication([])
window = UI()
window.showMaximized()
app.exec()